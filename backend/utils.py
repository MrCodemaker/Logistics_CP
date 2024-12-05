from pydoc import doc
import openpyxl
import pandas as pd
from docx import Document
from docx.shared import Inches, Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from logging import json
from pathlib import Path
from typing import Dict, Optional, Any, Tuple, List
from datetime import datetime
from flask import logging, render_template
from celery import shared_task
from .error_handler import ValidationError, FileError, log_error
from typing import List, Optional, Dict, Any
from pydantic import BaseModel, validator, ValidationError as PydanticValidationError
from openpyxl import load_workbook


# Добавляем кастомные исключения для обработки ошибок
class ExcelValidationError(Exception):
    # Кастомное исключение для ошибок валидации
    pass

class ExcelProcessingError(Exception):
    # Кастомное исключение для ошибок обработки
    pass

# Настраиваем систему логирования с ротацией файлов
def setup_logging():
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"app_{datetime.now().strftime('%Y%m%d')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler() # Также выводим в консоль
        ]
    )

"""
    Загружает конфигурацию из JSON-файла
    Args:
        config_path (str): Путь к файлу конфигурации
    
    Returns:
        Dict: Словарь с настройками
    
    Raises:
        FileNotFoundError: Если файл конфигурации не найден
        json.JSONDecodeError: Если файл содержит некорректный JSON
"""
def load_config(config_path: str) -> Dict:
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        logging.error(f"Файл конфигурации '{config_path}' не найден.")
        raise
    except json.JSONDecodeError:
        logging.error(f"Файл конфигурации '{config_path}' содержит некорректный JSON.")
        raise

"""
    Обрабатывает Excel-файл и извлекает данные согласно конфигурации
    
    Args:
        file_path (str): Путь к Excel-файлу
        config (Dict): Словарь с настройками извлечения данных
    
    Returns:
        Optional[Dict[str, Any]]: Словарь с извлеченными данными или None при ошибке
    """

class ExcelData(BaseModel):
    step_1: str
    step_2: str
    step_3: str
    step_1_deadline: str #  Пока str, позже изменим на datetime
    step_2_deadline: str #  Пока str, позже изменим на datetime
    step_3_deadline: str #  Пока str, позже изменим на datetime
    total_deadline: str  # Пока str, позже изменим на datetime
    tax: float
    total_tax: float

    @validator('tax', 'total_tax')
    def check_positive(cls, value):
        if value < 0:
            raise ValueError("Значение должно быть положительным")
        return value

    # Добавить валидаторы для дат позже, после выбора формата

def process_excel(file_path: str, config: Dict) -> Optional[Dict[str, Any]]:
    try:
        # Проверяем существование файла
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel файл не найден: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        data = {}
        
        # Перебираем все листы из конфигурации
        for sheet_config in config:
            sheet_name = sheet_config["sheet_name"]
            if sheet_name not in workbook.sheetnames:
                logging.warning(f"Лист '{sheet_name}' не найден")
                continue

            sheet = workbook[sheet_name]
            
            # Извлекаем данные согласно маппингу
            for key, cell_address in sheet_config["data_mapping"].items():
                try:
                    cell = sheet[cell_address]
                    # Добавляем базовую валидацию данных
                    if cell.value is not None:
                        data[key] = cell.value
                    else:
                        logging.warning(f"Пустая ячейка: {cell_address} на листе {sheet_name}")
                        data[key] = None
                except KeyError:
                    logging.error(f"Неверный адрес ячейки: {cell_address}")
                    data[key] = None

        return data if data else None

    except Exception as e:
        logging.exception("Ошибка при обработке Excel файла")
        return None
    
# Валидация данных
try:
    validated_data = ExcelData(**data)  # Валидация с помощью Pydantic
    return True, None, validated_data.dict()
except PydanticValidationError as e:
    errors = [f"{err['loc'][0]}: {err['msg']}" for err in e.errors()]
    return False, "\n".join(errors), None

except FileNotFoundError:
        return False, f"Файл не найден: {file_path}", None
    except Exception as e:
        # ... (логирование и обработка других ошибок)
        return False, str(e), None

"""
    Генерирует Word-документ на основе шаблона и данных
    
    Args:
        data (Dict[str, Any]): Данные для вставки в документ
        template_path (str): Путь к шаблону Word
        output_path (str): Путь для сохранения результата
    
    Returns:
        Optional[str]: Путь к созданному файлу или None при ошибке
    """
def generate_word(data: Dict[str, Any], template_path: str, output_path: str) -> Optional[str]:
    try:
        doc = Document(template_path)
        
        # Обрабатываем закладки в документе
        for bookmark in doc.bookmarks:
            if bookmark.name in data:
                value = data[bookmark.name]
                
                # Форматируем значение в зависимости от типа
                formatted_value = format_value(value)
                
                # Вставляем значение в закладку
                run = bookmark.parent.add_run()
                run.text = formatted_value
                
                # Применяем базовое форматирование
                if isinstance(value, (int, float)):
                    run.font.size = Pt(11)
                    run.font.name = 'Arial'

        # Сохраняем документ
        output_file = Path(output_path)
        output_file.parent.mkdir(exist_ok=True)
        doc.save(output_file)
        
        logging.info(f"Документ успешно создан: {output_file}")
        return str(output_file)

    except Exception as e:
        logging.exception("Ошибка при генерации Word документа")
        return None

"""
    Форматирует значение для вставки в документ
    
    Args:
        value (Any): Значение для форматирования
    
    Returns:
        str: Отформатированное значение
"""
def format_value(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:,.2f}".replace(',', ' ')
    elif isinstance(value, bool):
        return "Да" if value else "Нет"
    elif value is None:
        return ""
    else:
        return str(value)

def validate_excel_file(file):
    """
    Расширенная валидация Excel файла с подробным логированием
    
    Args:
        file: Загруженный файл
        
    Returns:
        Tuple[bool, str, Dict[str, Any]]: (успех, сообщение, дополнительные данные)
    """
    try:
        logging.info(f"Начало валидации файла: {file.filename}")
        
        # Проверка расширения файла
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise ValidationError(
                message = "Неверный формат файла. Допускаются только .xlsx или .xls"
                details={'filename': file.filename}
                logging.error(message)
                raise ExcelValidationError(message)

        # Проверка размера файла
        file_size = len(file.read())
        if file_size > 10 * 1024 * 1024:  # 10MB
            raise FileError(
                message="Файл слишком большой. Максимальный размер 10MB",
                details={'size': file_size, 'max_size': 10 * 1024 * 1024}
            )
        logging.info(f"Размер файла: {file_size / (1024*1024):.2f} MB")
        
        if file_size > 10 * 1024 * 1024:  # 10MB
            raise FileError(
                message = "Файл слишком большой. Максимальный размер 10MB"
                details={'size': file_size, 'max_size': 10 * 1024 * 1024}
                logging.error(message)
                raise ExcelValidationError(message)
            
        file.seek(0)

        # Чтение файла с помощью pandas
        df = pd.read_excel(file)
        logging.info(f"Файл успешно прочитан. Количество строк: {len(df)}")

        # Проверка структуры и данных
        validation_results = validate_excel_structure(df)
        if not validation_results['is_valid']:
            logging.error(f"Ошибка структуры: {validation_results['errors']}")
            raise ExcelValidationError('\n'.join(validation_results['errors']))

        # Валидация данных
        data_validation = validate_excel_data(df)
        if not data_validation['is_valid']:
            logging.error(f"Ошибка данных: {data_validation['errors']}")
            raise ExcelValidationError('\n'.join(data_validation['errors']))

        return True, "Файл успешно проверен", {
            'preview': df.head().to_dict(),
            'summary': {
                'total_rows': len(df),
                'total_sum': df['price'].sum() if 'price' in df.columns else 0
            }
        }
    
    except (ValidationError, FileError) as e:
        log_error(e, {'filename': file.filename})
        raise

    except Exception as e:
        log_error(e, {'filename': file.filename})
        raise FileError(f"Ошибка при обработке файла: {str(e)}")

        # Подготовка превью данных
        preview_data = {
            'preview': df.head().to_dict(),
            'summary': {
                'total_rows': len(df),
                'total_sum': df['price'].sum() if 'price' in df.columns else 0
            }
        }

        logging.info("Валидация успешно завершена")
        return True, "Файл успешно проверен", preview_data

    except ExcelValidationError as e:
        return False, str(e), {}
    except Exception as e:
        logging.exception("Непредвиденная ошибка при валидации")
        return False, f"Ошибка при обработке файла: {str(e)}", {}

def validate_excel_structure(df: pd.DataFrame) -> Dict[str, Any]:
    """Проверка структуры Excel файла"""
    required_columns = ['name', 'price', 'quantity']
    errors = []

    # Проверка наличия обязательных колонок
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        errors.append(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")

    # Дополнительные проверки структуры
    if len(df.columns) < len(required_columns):
        errors.append("Недостаточно колонок в файле")

    return {
        'is_valid': len(errors) == 0,
        'errors': errors
    }

def validate_excel_data(df: pd.DataFrame) -> Dict[str, Any]:
    """Расширенная валидация данных в Excel файле"""
    errors = []
    
    # Проверка цен
    if 'price' in df.columns:
        invalid_prices = df[~df['price'].apply(lambda x: isinstance(x, (int, float)) and x > 0)]
        if not invalid_prices.empty:
            errors.append(f"Неверные цены в строках: {invalid_prices.index.tolist()}")

    # Проверка количества
    if 'quantity' in df.columns:
        invalid_quantities = df[~df['quantity'].apply(lambda x: isinstance(x, (int, float)) and x > 0)]
        if not invalid_quantities.empty:
            errors.append(f"Неверное количество в строках: {invalid_quantities.index.tolist()}")

    # Проверка наименований
    if 'name' in df.columns:
        empty_names = df[df['name'].isna() | (df['name'] == '')]
        if not empty_names.empty:
            errors.append(f"Пустые наименования в строках: {empty_names.index.tolist()}")

    return {
        'is_valid': len(errors) == 0,
        'errors': errors
    }

@shared_task
def async_validate_excel_file(file_path: str) -> Dict[str, Any]:
    """
    Асинхронная валидация больших файлов через Celery
    
    Args:
        file_path: Путь к файлу для валидации
        
    Returns:
        Dict[str, Any]: Результат валидации
    """
    try:
        logging.info(f"Начало асинхронной валидации: {file_path}")
        
        # Чтение файла порциями для больших файлов
        chunk_size = 1000
        chunks = pd.read_excel(file_path, chunksize=chunk_size)
        
        all_errors = []
        total_rows = 0
        
        for chunk_num, chunk in enumerate(chunks, 1):
            validation_result = validate_excel_data(chunk)
            if not validation_result['is_valid']:
                all_errors.extend(validation_result['errors'])
            total_rows += len(chunk)
            logging.info(f"Обработано {chunk_num * chunk_size} строк")

        if all_errors:
            return {
                'status': 'error',
                'errors': all_errors,
                'total_rows': total_rows
            }

        return {
            'status': 'success',
            'message': 'Валидация успешно завершена',
            'total_rows': total_rows
        }

    except Exception as e:
        logging.exception("Ошибка при асинхронной валидации")
        return {
            'status': 'error',
            'message': str(e)
        }

# Инициализация при импорте модуля
setup_logging()